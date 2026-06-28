#!/usr/bin/env python3
"""Build the 'missing images to upload' sheet (CSV + XLSX) from layout.json."""
import json, os, csv

ROOT = os.path.dirname(os.path.abspath(__file__))
WORK = os.path.join(ROOT, "work")
L = json.load(open(os.path.join(WORK, "layout.json")))
M = json.load(open(os.path.join(WORK, "manifest.json")))
OLD_BASE = "https://img.damensch.com/products-old"
CSV_OUT = os.path.join(ROOT, "damensch_missing_images_upload.csv")
XLSX_OUT = os.path.join(ROOT, "damensch_missing_images_upload.xlsx")

def old_src(pkey, order):
    for it in M[pkey]["old"]:
        if it["order"] == order:
            return it["src"]
    return ""

HEADERS = ["new pdp url", "old image url", "display order", "image type", "old position", "product"]

rows = []
for pkey in ["pdp1", "pdp2"]:
    d = L[pkey]
    url = d["url"]; title = d["title"]; new_count = d["new_count"]
    # missing = columns with new_order None, kept in old display order
    missing = [c for c in d["columns"] if c["new_order"] is None]
    missing.sort(key=lambda c: c["old_order"])
    pos = new_count + 1
    for c in missing:
        rows.append({
            "new pdp url": url,
            "old image url": f'{OLD_BASE}/{old_src(pkey, c["old_order"])}',
            "display order": pos,
            "image type": c["old_label"],
            "old position": c["old_order"],
            "product": title,
        })
        pos += 1

# ---- CSV ----
with open(CSV_OUT, "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=HEADERS)
    w.writeheader()
    w.writerows(rows)
print("wrote", os.path.basename(CSV_OUT), f"({len(rows)} rows)")

# ---- XLSX ----
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook(); ws = wb.active; ws.title = "to_upload"
    head_fill = PatternFill("solid", fgColor="101828")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    link_font = Font(color="1570EF", underline="single")
    thin = Side(style="thin", color="EAECF0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = head_fill; cell.font = head_font
        cell.alignment = Alignment(vertical="center"); cell.border = border

    pdp_blocks = {}
    for ri, r in enumerate(rows, 2):
        for ci, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=ri, column=ci, value=r[h])
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(h in ("image type", "product")))
        # hyperlinks
        c_pdp = ws.cell(row=ri, column=1); c_pdp.hyperlink = r["new pdp url"]; c_pdp.font = link_font
        c_img = ws.cell(row=ri, column=2); c_img.hyperlink = r["old image url"]; c_img.font = link_font
        pdp_blocks.setdefault(r["new pdp url"], []).append(ri)

    widths = {"new pdp url": 52, "old image url": 64, "display order": 13, "image type": 34, "old position": 12, "product": 34}
    for ci, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths[h]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows)+1}"

    wb.save(XLSX_OUT)
    print("wrote", os.path.basename(XLSX_OUT))
except Exception as e:
    print("xlsx skipped:", e)
