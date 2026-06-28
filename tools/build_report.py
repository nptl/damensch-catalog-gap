#!/usr/bin/env python3
"""Reconcile analyst+skeptic, run consistency checks, build 90-PDP report + Excel upload sheet."""
import json, os, html, csv

ROOT = os.path.dirname(os.path.abspath(__file__))
WORK = os.path.join(ROOT, "work")
man = {p["key"]: p for p in json.load(open(os.path.join(WORK, "pdps_manifest.json")))["pdps"]}
order = [p["key"] for p in json.load(open(os.path.join(WORK, "pdps_manifest.json")))["pdps"]]
A = json.load(open(os.path.join(WORK, "analysis90.json")))
OVR_PATH = os.path.join(WORK, "match_overrides.json")
OVERRIDES = json.load(open(OVR_PATH)) if os.path.exists(OVR_PATH) else {}  # {key:[old_seq,...]} matches to break
RVW_PATH = os.path.join(WORK, "review_flags.json")
REVIEW = set(json.load(open(RVW_PATH))) if os.path.exists(RVW_PATH) else set()
HTML_OUT = os.path.join(ROOT, "damensch_catalog_comparison_90.html")
XLSX_OUT = os.path.join(ROOT, "damensch_missing_images_upload_90.xlsx")
CSV_OUT  = os.path.join(ROOT, "damensch_missing_images_upload_90.csv")

def esc(s): return html.escape(str(s if s is not None else ""))

def reconcile(key):
    p = man[key]; rec = A.get(key) or {}
    a = rec.get("analyst"); sk = rec.get("skeptic")
    olds = p["old"]; news = p["new"]; M = len(news)
    flags = []
    lab_old = {}; lab_new = {}
    if not a:
        flags.append("ANALYST_FAILED")
        return {"p": p, "olds": olds, "news": news, "align": {o["seq"]: None for o in olds},
                "lab_old": {o["seq"]: ("image","other") for o in olds},
                "lab_new": {n["idx"]: ("image","other") for n in news},
                "missing": [o["seq"] for o in olds], "extra": list(range(1,M+1)),
                "matched": 0, "flags": flags}
    for c in a.get("old", []): lab_old[c["seq"]] = (c.get("label","image"), c.get("type","other"))
    for c in a.get("new", []): lab_new[c["idx"]] = (c.get("label","image"), c.get("type","other"))
    for o in olds: lab_old.setdefault(o["seq"], ("image","other"))
    for n in news: lab_new.setdefault(n["idx"], ("image","other"))

    align = {}; seen=set(); used=set(); reuse=0
    for it in a.get("alignment", []):
        s = it.get("old_seq"); ni = it.get("new_idx")
        if s in seen or s is None: continue
        seen.add(s)
        if ni is not None and (ni in used or ni < 1 or ni > M):
            if ni in used: reuse += 1
            ni = None
        if ni is not None: used.add(ni)
        align[s] = ni
    for o in olds:
        if o["seq"] not in align:
            align[o["seq"]] = None; flags.append(f"old{o['seq']}_unmapped")
    if reuse: flags.append(f"reuse_blocked:{reuse}")
    # skeptic rescue
    if sk and sk.get("reviewed"):
        for rv in sk["reviewed"]:
            s = rv.get("old_seq")
            if s in align and align[s] is None and not rv.get("still_missing"):
                ni = rv.get("matched_new_idx")
                if isinstance(ni,int) and 1 <= ni <= M and ni not in used:
                    align[s] = ni; used.add(ni)
    # break verified-wrong matches (from match-verification pass)
    broken = OVERRIDES.get(key, [])
    for s in broken:
        if s in align and align[s] is not None:
            align[s] = None
    if key in REVIEW: flags.append("manual_review")
    # recompute from final alignment
    used = set(v for v in align.values() if v is not None)
    missing = sorted([s for s,ni in align.items() if ni is None])
    extra = [i for i in range(1, M+1) if i not in used]
    return {"p": p, "olds": olds, "news": news, "align": align, "lab_old": lab_old,
            "lab_new": lab_new, "missing": missing, "extra": extra, "matched": len(used), "flags": flags}

R = {k: reconcile(k) for k in order}

# ---------------- consistency checks ----------------
problems = []
for k in order:
    r = R[k]; M = len(r["news"]); O = len(r["olds"])
    matched = r["matched"]; miss = len(r["missing"]); extra = len(r["extra"])
    if matched + miss != O: problems.append((k, f"old reconcile {matched}+{miss}!={O}"))
    if matched + extra != M: problems.append((k, f"new reconcile {matched}+{extra}!={M}"))
    if len(set(r["align"].keys())) != O: problems.append((k, "old key count"))
    if r["flags"]: problems.append((k, ";".join(r["flags"])))

# ---------------- Excel / CSV upload sheet ----------------
OLD_W = "https://img.damensch.com/products-old/"  # not used directly; we use manifest urls
HEADERS = ["new pdp url","old image url","display order","image type","old position","product"]
rows = []
for k in order:
    r = R[k]; p = r["p"]; M = len(r["news"]); pos = M + 1
    omap = {o["seq"]: o for o in r["olds"]}
    for s in r["missing"]:
        o = omap[s]; lab, typ = r["lab_old"][s]
        rows.append({"new pdp url": p["url"], "old image url": o["url"], "display order": pos,
                     "image type": lab, "old position": s, "product": p["title"]})
        pos += 1
with open(CSV_OUT,"w",newline="") as f:
    w=csv.DictWriter(f,fieldnames=HEADERS); w.writeheader(); w.writerows(rows)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb=Workbook(); ws=wb.active; ws.title="to_upload"
    hf=PatternFill("solid",fgColor="101828"); hfont=Font(color="FFFFFF",bold=True)
    lf=Font(color="1570EF",underline="single"); thin=Side(style="thin",color="EAECF0"); bd=Border(thin,thin,thin,thin)
    for ci,h in enumerate(HEADERS,1):
        c=ws.cell(1,ci,h); c.fill=hf; c.font=hfont; c.border=bd
    for ri,r in enumerate(rows,2):
        for ci,h in enumerate(HEADERS,1):
            c=ws.cell(ri,ci,r[h]); c.border=bd
            c.alignment=Alignment(vertical="center",wrap_text=(h in("image type","product","new pdp url")))
        ws.cell(ri,1).hyperlink=r["new pdp url"]; ws.cell(ri,1).font=lf
        ws.cell(ri,2).hyperlink=r["old image url"]; ws.cell(ri,2).font=lf
    for ci,h in enumerate(HEADERS,1):
        ws.column_dimensions[get_column_letter(ci)].width={"new pdp url":48,"old image url":60,"display order":13,"image type":30,"old position":12,"product":34}[h]
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:F{len(rows)+1}"
    # summary tab
    s2=wb.create_sheet("summary")
    s2.append(["product","new pdp url","old count","new count","matched","missing","flags"])
    for cell in s2[1]: cell.font=Font(bold=True)
    for k in order:
        r=R[k]; s2.append([r["p"]["title"], r["p"]["url"], len(r["olds"]), len(r["news"]), r["matched"], len(r["missing"]), ";".join(r["flags"])])
    for col,wd in zip("ABCDEFG",[34,48,10,10,9,9,26]): s2.column_dimensions[col].width=wd
    s2.freeze_panes="A2"
    wb.save(XLSX_OUT)
    xlsx_ok=True
except Exception as e:
    xlsx_ok=str(e)

# ---------------- HTML report ----------------
def old_cell(r,o):
    seq=o["seq"]; lab,_=r["lab_old"][seq]
    return f'''<div class="ocell"><div class="imw"><span class="ob old">{seq}</span>
      <img loading="lazy" class="oimg" src="{esc(o["url"])}"></div>
      <div class="olabel">{esc(lab)}</div></div>'''
def new_cell(r,ni):
    if ni is None: return '<div class="ncell"><div class="na">Not Available</div></div>'
    n=next(x for x in r["news"] if x["idx"]==ni)
    return f'''<div class="ncell"><div class="imw"><span class="ob new">{ni}</span>
      <img loading="lazy" class="nimg" src="{esc(n["url"])}"></div></div>'''
def extra_col(r,idx):
    n=next(x for x in r["news"] if x["idx"]==idx); lab,_=r["lab_new"][idx]
    return f'''<div class="col extra"><div class="ocell"><div class="imw blank"><span class="addtag">extra new</span></div>
      <div class="olabel muted">{esc(lab)}</div></div>
      <div class="ncell"><div class="imw"><span class="ob new">{idx}</span>
      <img loading="lazy" class="nimg" src="{esc(n["url"])}"></div></div></div>'''
def proj(r):
    M=len(r["news"]); cards=[]
    for n in r["news"]:
        cards.append(f'''<div class="pcard"><div class="pimw"><span class="pb cur">{n["idx"]}</span>
          <img loading="lazy" class="pimg cur" src="{esc(n["url"])}"></div><div class="ptag cur">current 3:4</div></div>''')
    pos=M+1; omap={o["seq"]:o for o in r["olds"]}
    for s in r["missing"]:
        o=omap[s]; lab,_=r["lab_old"][s]
        cards.append(f'''<div class="pcard"><div class="pimw"><span class="pb add">{pos}</span><span class="newflag">NEW</span>
          <img loading="lazy" class="pimg add" src="{esc(o["url"])}"></div><div class="ptag add">added 9:16</div>
          <div class="pdesc">{esc(lab)}</div></div>'''); pos+=1
    return "".join(cards)

def section(k):
    r=R[k]; p=r["p"]; O=len(r["olds"]); M=len(r["news"]); miss=len(r["missing"])
    cols="".join(f'<div class="col">{old_cell(r,o)}{new_cell(r,r["align"][o["seq"]])}</div>' for o in r["olds"])
    cols+="".join(extra_col(r,i) for i in r["extra"])
    dcls="neg" if M<O else ("pos" if M>O else "zero")
    dp = f"{O-M} fewer" if M<O else (f"{M-O} more" if M>O else "same count")
    flagbadge = f'<span class="flag">⚑ review</span>' if r["flags"] else ""
    return f'''<section class="pdp" id="{esc(k)}">
      <a class="title" href="{esc(p["url"])}" target="_blank">{esc(p["title"])}</a>{flagbadge}
      <div class="count">{O} <span class="ar">&rarr;</span> {M} images
        <span class="badge {dcls}">{dp}</span>
        <span class="mm">{r["matched"]} matched &middot; <b class="miss">{miss} missing</b></span></div>
      <div class="rowlabel">Old (9:16) vs current new (3:4)</div>
      <div class="carousel">{cols}</div>
      <div class="proj-head"><div class="rowlabel proj">After upload — projected new gallery</div></div>
      <div class="proj">{proj(r)}</div>
    </section>'''

# master index
tot_old=sum(len(R[k]["olds"]) for k in order); tot_new=sum(len(R[k]["news"]) for k in order)
tot_miss=sum(len(R[k]["missing"]) for k in order)
idx_rows=""
for k in order:
    r=R[k]; O=len(r["olds"]); M=len(r["news"]); miss=len(r["missing"])
    cls="hi" if miss>=5 else ("md" if miss>=1 else "lo")
    fl="⚑" if r["flags"] else ""
    idx_rows+=f'<tr class="{cls}"><td><a href="#{esc(k)}">{esc(r["p"]["title"])}</a> {fl}</td><td>{O}</td><td>{M}</td><td>{r["matched"]}</td><td class="mc">{miss}</td></tr>'

sections="".join(section(k) for k in order)
doc=f'''<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>DaMENSCH — Catalog Gap (90 PDPs)</title>
<style>
*{{box-sizing:border-box}} body{{margin:0;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;color:#101828;background:#f8f9fb}}
.wrap{{padding:18px 16px 64px;max-width:1400px;margin:0 auto}}
h1{{font-size:20px;margin:0 0 2px}} .sub{{font-size:12px;color:#98a2b3;margin-bottom:14px}}
.stats{{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:16px}}
.stat{{background:#fff;border:1px solid #eaecf0;border-radius:10px;padding:8px 12px}} .stat .n{{font-size:18px;font-weight:700}} .stat .l{{font-size:11px;color:#667085}}
.stat.bad .n{{color:#b42318}}
table.idx{{width:100%;border-collapse:collapse;background:#fff;border:1px solid #eaecf0;border-radius:10px;overflow:hidden;font-size:13px;margin-bottom:20px}}
.idx th,.idx td{{padding:6px 10px;border-bottom:1px solid #f2f4f7;text-align:left}}
.idx th{{background:#f8f9fb;font-size:11px;text-transform:uppercase;color:#667085}}
.idx td:nth-child(n+2){{text-align:center;width:70px}} .idx a{{color:#101828;text-decoration:none}} .idx a:hover{{color:#1570ef}}
.idx td.mc{{font-weight:700}} .idx tr.hi td.mc{{color:#b42318}} .idx tr.md td.mc{{color:#b54708}} .idx tr.lo td.mc{{color:#067647}}
section.pdp{{background:#fff;border:1px solid #eaecf0;border-radius:14px;padding:16px;margin:14px 0;scroll-margin-top:10px}}
.title{{font-size:17px;font-weight:700;color:#101828;text-decoration:none}} .title:hover{{color:#1570ef;text-decoration:underline}}
.flag{{font-size:11px;color:#b54708;background:#fffaeb;border:1px solid #fec84b;border-radius:6px;padding:2px 6px;margin-left:8px}}
.count{{font-size:13px;color:#475467;margin:6px 0 12px}} .ar{{color:#98a2b3}}
.badge{{font-size:12px;font-weight:600;padding:3px 9px;border-radius:999px;margin-left:8px}}
.badge.neg{{background:#fef3f2;color:#b42318}} .badge.pos{{background:#ecfdf3;color:#067647}} .badge.zero{{background:#f2f4f7;color:#475467}}
.mm{{font-size:12px;color:#667085;margin-left:8px}} .mm .miss{{color:#b42318}}
.rowlabel{{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.04em;color:#667085;margin:4px 0 8px}} .rowlabel.proj{{color:#b54708}}
.carousel,.proj{{display:flex;gap:12px;overflow-x:auto;padding-bottom:10px}}
.col{{flex:0 0 170px;width:170px}} .imw{{position:relative}}
.oimg{{width:100%;aspect-ratio:9/16;object-fit:cover;border-radius:8px;border:1px solid #eaecf0;display:block;background:#f2f4f7}}
.nimg{{width:100%;aspect-ratio:3/4;object-fit:cover;border-radius:8px;border:1px solid #d1e9ff;display:block;background:#f2f4f7}}
.imw.blank{{width:100%;aspect-ratio:9/16;border:1.5px dashed #d0d5dd;border-radius:8px;background:#f9fafb;display:flex;align-items:center;justify-content:center}}
.addtag{{font-size:11px;color:#98a2b3;font-weight:600}}
.ob{{position:absolute;top:5px;left:5px;color:#fff;font-size:11px;font-weight:700;min-width:20px;height:20px;padding:0 5px;border-radius:10px;display:flex;align-items:center;justify-content:center}}
.ob.old{{background:rgba(16,24,40,.82)}} .ob.new{{background:#1570ef}}
.olabel{{font-size:11px;font-weight:600;line-height:1.2;margin:6px 0 8px;min-height:40px;display:-webkit-box;-webkit-line-clamp:3;-webkit-box-orient:vertical;overflow:hidden}}
.olabel.muted{{color:#98a2b3;font-weight:500}}
.na{{width:100%;aspect-ratio:3/4;display:flex;align-items:center;justify-content:center;border:1.5px dashed #fda29b;background:#fef3f2;color:#b42318;font-weight:700;font-size:12px;border-radius:8px}}
.proj-head{{margin-top:16px;border-top:1px solid #eaecf0;padding-top:12px}}
.proj{{align-items:flex-start}} .pcard{{flex:0 0 130px;width:130px}} .pimw{{position:relative}}
.pimg{{width:100%;object-fit:cover;border-radius:8px;display:block;background:#f2f4f7}}
.pimg.cur{{aspect-ratio:3/4;border:1px solid #d1e9ff}} .pimg.add{{aspect-ratio:9/16;border:2px solid #f79009}}
.pb{{position:absolute;top:5px;left:5px;color:#fff;font-size:11px;font-weight:700;min-width:20px;height:20px;padding:0 5px;border-radius:10px;display:flex;align-items:center;justify-content:center}}
.pb.cur{{background:#1570ef}} .pb.add{{background:#f79009}}
.newflag{{position:absolute;top:5px;right:5px;background:#f79009;color:#fff;font-size:8px;font-weight:800;padding:2px 4px;border-radius:4px}}
.ptag{{font-size:10px;font-weight:700;margin-top:5px}} .ptag.cur{{color:#1570ef}} .ptag.add{{color:#b54708}}
.pdesc{{font-size:10px;color:#667085;line-height:1.15;margin-top:2px;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;overflow:hidden}}
</style></head><body><div class="wrap">
<h1>DaMENSCH — Catalog Image Gap (90 PDPs)</h1>
<div class="sub">old website (9:16) vs new website (3:4) · row 3 = projected new gallery after uploading the missing old images · images from img.damensch.com</div>
<div class="stats">
  <div class="stat"><div class="n">90</div><div class="l">PDPs</div></div>
  <div class="stat"><div class="n">{tot_old} &rarr; {tot_new}</div><div class="l">old &rarr; new images</div></div>
  <div class="stat bad"><div class="n">{tot_miss}</div><div class="l">missing images to upload</div></div>
</div>
<table class="idx"><thead><tr><th>Product</th><th>Old</th><th>New</th><th>Matched</th><th>Missing</th></tr></thead><tbody>{idx_rows}</tbody></table>
{sections}
</div></body></html>'''
open(HTML_OUT,"w").write(doc)

print(f"PDPs: {len(order)} | old {tot_old} new {tot_new} | missing total {tot_miss}")
print(f"report: {os.path.basename(HTML_OUT)} ({round(os.path.getsize(HTML_OUT)/1024,1)} KB)")
print(f"xlsx: {xlsx_ok}  rows: {len(rows)}")
print(f"consistency problems: {len(problems)}")
for k,m in problems[:25]: print("  ", k, "->", m)
